# xlsx_writer.py
from openpyxl import load_workbook, Workbook
import utility.config as config
from utility.logger import logger
import os

class XLSXWriter(object):
    def __init__(self, file_name: str=config.LOCAL_FILE_NAME ) -> None:
        self.__file_name = file_name

    async def write_sensor_data(self, data: dict) -> None:
        """
            Method for writing data to local XLSX file
        """
        if os.path.exists(self.__file_name):
            logger.info(f"Opening workbook named {self.__file_name}")
            workbook = load_workbook(self.__file_name)
        else:
            logger.warning(f"Workbork {self.__file_name} does not exist. Creating new one.")
            workbook = Workbook()
        
        for sensor_name, sensor_data in data.items():
            if sensor_name not in workbook.sheetnames:
                workbook.create_sheet(sensor_name)

            sheet = workbook[sensor_name]

            if not sheet.iter_rows(min_row=1, max_row=1, max_col=1):
                headers = list(sensor_data.keys())
                sheet.append(headers)

            data_row = [sensor_data[key] for key in headers]
            sheet.append(data_row)

        workbook.save(self.__file_name)